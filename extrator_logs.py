from flask import Flask, render_template, request, redirect, render_template_string, jsonify, send_from_directory, send_file, Blueprint
from version import __version__
import os, sys, random, string, json, re, smtplib, webbrowser, zipfile, logging, threading, uuid, mimetypes
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import oracledb, csv, subprocess
from datetime import datetime
from logger import logger, USUARIO_WINDOWS
import execucao
from execucao import registrar_execucao, gerar_pid
from atualizacao_agente import obter_agentes, enviar_atualizacao, TAMANHO_MAXIMO_MB
from parametrizacao_pdv import pagina_configurar_pdv, enviar_configuracao_pdv, verificar_configuracao_pdv, relatorio_parametrizacao, versao_pdv, reiniciar_pdv, fechar_pdv
from request_api import (pagina_requisicao_api, fazer_requisicao_api, obter_apis,
                         gerar_retorno_download, baixar_retorno, enviar_retorno_email)
from mdm import pagina_mdm, consultar_cliente_mdm, cadastrar_cliente_mdm, atualizar_cliente_mdm

# BUNDLE_DIR: recursos somente-leitura empacotados (templates, static)
# APP_DIR:    pasta do .exe / pasta do script — para arquivos editáveis (config, output, log)
if getattr(sys, 'frozen', False):
    BUNDLE_DIR = sys._MEIPASS
    APP_DIR    = os.path.dirname(sys.executable)
else:
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    APP_DIR    = BUNDLE_DIR

app = Flask(__name__,
            template_folder=os.path.join(BUNDLE_DIR, 'templates'),
            static_folder=os.path.join(BUNDLE_DIR, 'static'))

# Teto do upload do pacote de atualização do agente (aba Administrador).
# Uma folga sobre o limite do anexo permite que a validação devolva uma mensagem
# clara em vez de o Flask cortar a requisição com 413.
app.config['MAX_CONTENT_LENGTH'] = (TAMANHO_MAXIMO_MB + 5) * 1024 * 1024

_img_bp = Blueprint('img', __name__,
                    static_folder=os.path.join(APP_DIR, 'img'),
                    static_url_path='/img')
app.register_blueprint(_img_bp)

PROPS_DIR   = os.path.join(APP_DIR, 'properties')
CONFIG_FILE = os.path.join(PROPS_DIR, 'config.properties')
SECURE_FILE = os.path.join(PROPS_DIR, 'secure.properties')
OUTPUT_DIR  = os.path.join(APP_DIR, 'output')

# Werkzeug loga uma linha por requisição HTTP (incluindo cada CSS/imagem), o que
# domina o arquivo de log. Só acompanha o nível da aplicação quando em DEBUG;
# fora disso registra apenas erros.
_wz = logging.getLogger('werkzeug')
_wz.setLevel(logging.INFO if logger.level <= logging.DEBUG else logging.ERROR)

# Redireciona werkzeug para o mesmo arquivo de log quando empacotado
if getattr(sys, 'frozen', False):
    from logger import _LOG_PATH
    for _h in logger.handlers:
        _wz.addHandler(_h)
        logging.getLogger('flask.app').addHandler(_h)
    logging.getLogger('flask.app').setLevel(logger.level)
    # Redireciona stdout/stderr para o log (sem console visível)
    sys.stdout = open(_LOG_PATH, 'a', encoding='utf-8')
    sys.stderr = sys.stdout

logger.info(f"===== Backoffice Equipe QA v{__version__} iniciado =====")

@app.context_processor
def inject_version():
    return {'app_version': __version__}


def ler_properties(arquivo):
    props = {}
    if not os.path.exists(arquivo):
        logger.warning(f"Arquivo de configuração não encontrado: {arquivo}")
        return {}
    try:
        with open(arquivo, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    key, value = line.split('=', 1)
                    props[key.strip()] = value.strip()  # SEMPRE salva como STRING
        return props
    except Exception as e:
        logger.error(f"Erro ao ler arquivo {arquivo}: {str(e)}")
        return {}

def ler_config_completo(*args):
    """Retorna config.properties + secure.properties mesclados (secure tem precedência).
    Aceita argumentos opcionais para ser compatível com a assinatura de ler_properties."""
    props = ler_properties(CONFIG_FILE)
    props.update(ler_properties(SECURE_FILE))
    return props

def get_oracle_conn(props):
    dsn = f"{props['oracle_host']}:{props['oracle_port']}/{props['oracle_service']}"
    try:
        conn = oracledb.connect(
            user=props['oracle_user'],
            password=props['oracle_password'],
            dsn=dsn
        )
        logger.debug(f"Conexão Oracle estabelecida: {dsn} (user={props['oracle_user']})")
        return conn
    except Exception as e:
        logger.error(f"Erro ao conectar ao Oracle: {str(e)}")
        raise

PROP_SECTIONS = [
    ("# === Log - nível de detalhe e rotação do arquivo log/extrator_logs.log ===\n"
     "# log.level: DEBUG | INFO | WARNING | ERROR | CRITICAL",
     lambda k: k.startswith('log.')),
    ("# === Registro de execução (e-mail por funcionalidade; true/false) ===",
     lambda k: k == 'registro_execucao'),
    ("# === Administrador - exibe a aba com as configurações sensíveis (true/false) ===",
     lambda k: k == 'admin_habilitado'),
    ("# === Atualização de agentes - assunto do e-mail que leva o pacote ===",
     lambda k: k == 'agentes_order' or k.startswith('agente.')),
    ("# === Abas - controla quais abas ficam visíveis na interface (true/false) ===",
     lambda k: k.startswith('tab.')),
    ("# === E-mails - lista de destinatários disponíveis ===",
     lambda k: k == 'emails_destino'),
    ("# === Modo de instalação e comunicação ===",
     lambda k: k in ('modo_instalacao', 'bec_loja', 'bec_pdv',
                     'pinpad_porta', 'pinpad_modo_comunicacao',
                     'logs_modo_comunicacao', 'pdv_modo_comunicacao',
                     'registro_modo_comunicacao', 'atualizacao_modo_comunicacao',
                     'bec_tunnel_url', 'pinpad_tunnel_token', 'ignorar_lojas')),
    ("# === Lojas e PDVs disponíveis para seleção ===",
     lambda k: k == 'stores' or k.endswith('_pdvs')),
    ("# === Logs disponíveis para solicitação ===",
     lambda k: k == 'logs'),
    ("# === Parâmetros de configuração de PDV ===",
     lambda k: k == 'PARAMETROS_PDV'),
    ("# === APIs externas - metadados (a APIKEY fica em secure.properties) ===",
     lambda k: k.startswith('api.') or k == 'api_order'),
    ("# === Oracle - consultas SQL disponíveis ===",
     lambda k: k.startswith('oracle_query')),
]

def salvar_properties(arquivo, props):
    try:
        escritas = set()
        linhas = []
        for comentario, pertence in PROP_SECTIONS:
            chaves_secao = [k for k in props if pertence(k)]
            if not chaves_secao:
                continue
            linhas.append(f"{comentario}\n")
            for key in chaves_secao:
                value = props[key]
                if isinstance(value, list):
                    value = ','.join(str(v) for v in value)
                elif isinstance(value, str):
                    value = value.strip()
                else:
                    value = str(value)
                linhas.append(f"{key}={value}\n")
                escritas.add(key)
            linhas.append("\n")

        # Chaves que não se encaixam em nenhuma seção
        restantes = [k for k in props if k not in escritas]
        if restantes:
            linhas.append("# === Outras configurações ===\n")
            for key in restantes:
                value = props[key]
                if isinstance(value, list):
                    value = ','.join(str(v) for v in value)
                else:
                    value = str(value).strip()
                linhas.append(f"{key}={value}\n")

        with open(arquivo, 'w', encoding='utf-8') as f:
            f.writelines(linhas)
        logger.info(f"Arquivo {arquivo} salvo com sucesso")
    except Exception as e:
        logger.error(f"Erro ao salvar arquivo {arquivo}: {str(e)}")

# Propriedades introduzidas por novas versões do BEC e seus valores padrão.
# O instalador usa "onlyifdoesntexist" no config.properties, então ao atualizar a
# aplicação o arquivo do usuário é preservado — e ficaria sem as chaves novas.
# Esta migração roda no start: cadastra o que falta com o padrão e nunca
# sobrescreve um valor já existente.
PROPS_PADRAO = {
    'admin_habilitado': 'false',
    'registro_execucao': 'true',
    'agentes_order': 'extrator,sp',
    'agente.extrator.nome': 'Agent Extrator Log',
    'agente.extrator.assunto': '[Atualizacao Agente]',
    'agente.sp.nome': 'Server Agent SP',
    'agente.sp.assunto': '[Atualizacao Agente SP]',
    'log.level': 'INFO',
    'log.console_level': 'INFO',
    'log.max_size_mb': '5',
    'log.backup_count': '3',
}


def garantir_props_padrao():
    """Cadastra no config.properties as chaves novas que ainda não existem."""
    props = ler_properties(CONFIG_FILE)
    if not props:
        return  # arquivo ausente/ilegível — nada a migrar
    faltando = {k: v for k, v in PROPS_PADRAO.items() if k not in props}
    if not faltando:
        return
    props.update(faltando)
    salvar_properties(CONFIG_FILE, props)
    logger.info(f"Propriedades novas cadastradas com valor padrão: {', '.join(sorted(faltando))}")


def salvar_secure_campos(arquivo, valores):
    """Atualiza chaves específicas do secure.properties preservando o restante.

    Mantém comentários, ordem e as demais chaves (inclusive as api.*.apikey, que
    têm gravação própria). Chaves ainda inexistentes são acrescentadas ao final.
    Valor None significa "manter o que já está gravado" — usado nos campos
    sensíveis que a tela exibe em branco.
    """
    try:
        pendentes = {k: v for k, v in valores.items() if v is not None}
        linhas = []
        if os.path.exists(arquivo):
            with open(arquivo, 'r', encoding='utf-8') as f:
                for line in f:
                    linha = line.rstrip('\n')
                    stripped = linha.strip()
                    if stripped and not stripped.startswith('#') and '=' in stripped:
                        chave = stripped.split('=', 1)[0].strip()
                        if chave in pendentes:
                            linhas.append(f'{chave}={pendentes.pop(chave)}')
                            continue
                    linhas.append(linha)

        if pendentes:
            while linhas and linhas[-1].strip() == '':
                linhas.pop()
            linhas.append('')
            for chave in sorted(pendentes):
                linhas.append(f'{chave}={pendentes[chave]}')

        with open(arquivo, 'w', encoding='utf-8') as f:
            f.write('\n'.join(linhas) + '\n')
        logger.info(f"Configurações sensíveis atualizadas ({len(valores)} campo(s))")
    except Exception as e:
        logger.error(f"Erro ao salvar {arquivo}: {str(e)}")
        raise


garantir_props_padrao()


def converter_data_para_oracle(data_ddmmyyyy):
    """Converte DD/MM/YYYY para YYYYMMDD"""
    if not data_ddmmyyyy:
        return ''
    try:
        dia, mes, ano = data_ddmmyyyy.split('/')
        data_oracle = f"{ano}{mes}{dia}"
        return data_oracle
    except Exception as e:
        logger.error(f"Erro ao converter data {data_ddmmyyyy}: {str(e)}")
        return data_ddmmyyyy

def exportar_consulta_para_csv(nome_consulta, loja='', pdv='', nsu='', data='', formato='csv', separador='.'):
    props = ler_config_completo()
    sql = props.get(f'oracle_query.{nome_consulta}')
    if not sql:
        erro = f"Consulta '{nome_consulta}' não encontrada no config.properties"
        logger.error(erro)
        raise ValueError(erro)

    logger.info(f"Exportando {nome_consulta} [loja={loja} pdv={pdv} nsu={nsu} data={data} formato={formato}]")

    # Converter data de DD/MM/YYYY para YYYYMMDD
    data_oracle = converter_data_para_oracle(data)

    # Substituir variáveis
    # $DATA_BR precisa ser substituído antes de $DATA (senão o replace de $DATA
    # corromperia o prefixo de $DATA_BR). Use $DATA para colunas numéricas/YYYYMMDD
    # (ex.: P2K_*) e $DATA_BR para colunas armazenadas como texto 'DD/MM/YYYY'.
    sql = sql.replace('$LOJA', loja if loja else '0')
    sql = sql.replace('$PDV', pdv if pdv else '0')
    sql = sql.replace('$NSU', nsu if nsu else '')
    sql = sql.replace('$DATA_BR', data if data else '')
    sql = sql.replace('$DATA', data_oracle if data_oracle else '')
    
    logger.debug(f"SQL final da consulta {nome_consulta}: {sql}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    
    # Definir extensão baseado no formato
    extensao = 'xlsx' if formato == 'xlsx' else 'csv'
    caminho_arquivo = os.path.join(OUTPUT_DIR, f'{nome_consulta}_{ts}.{extensao}')

    try:
        conn = get_oracle_conn(props)
        cur = conn.cursor()
        cur.execute(sql)
        colunas = [d[0] for d in cur.description]
        dados = cur.fetchall()
        
        if formato == 'xlsx':
            # Exportar para XLSX
            exportar_para_xlsx(caminho_arquivo, colunas, dados, separador)
        else:
            # Exportar para CSV
            exportar_para_csv(caminho_arquivo, colunas, dados, separador)
        
        linhas = len(dados)
        cur.close()
        conn.close()

        logger.info(f"Exportação concluída: {caminho_arquivo} ({linhas} linhas)")
        return caminho_arquivo, linhas
    except Exception as e:
        logger.error(f"Erro durante exportação de {nome_consulta}: {str(e)}")
        raise

def exportar_para_csv(caminho, colunas, dados, separador='.'):
    """Exporta dados para CSV com separador decimal configurável e suporte a LOBs"""
    try:
        with open(caminho, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';', lineterminator='\n')
            writer.writerow(colunas)

            erros_lob = 0
            for row in dados:
                row_processada = []
                for valor in row:
                    # Converter LOBs
                    if hasattr(valor, 'read'):  # LOB object
                        try:
                            valor = valor.read()
                        except Exception as e:
                            # Só a primeira falha é logada — evita uma linha por registro
                            if not erros_lob:
                                logger.warning(f"Erro ao ler LOB: {str(e)}")
                            erros_lob += 1
                            valor = "[LOB não lido]"
                    
                    if isinstance(valor, bytes):
                        # Se for bytes, tenta decodificar
                        try:
                            valor = valor.decode('utf-8')
                        except:
                            valor = "[Dados binários]"
                    
                    if isinstance(valor, float):
                        # Formatar número com separador decimal escolhido
                        valor_str = str(valor).replace('.', ',') if separador == ',' else str(valor)
                        row_processada.append(valor_str)
                    else:
                        row_processada.append(str(valor) if valor is not None else '')
                writer.writerow(row_processada)

        if erros_lob > 1:
            logger.warning(f"{erros_lob} LOBs não puderam ser lidos em {os.path.basename(caminho)}")
    except Exception as e:
        logger.error(f"Erro ao exportar para CSV: {str(e)}")
        raise


def exportar_para_xlsx(caminho, colunas, dados, separador='.'):
    """Exporta dados para XLSX com separador decimal configurável e suporte a LOBs"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        # Adicionar header com estilo
        header_fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, coluna in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = coluna
            cell.fill = header_fill
            cell.font = header_font
        
        # Adicionar dados
        erros_lob = 0
        for row_num, row in enumerate(dados, 2):
            for col_num, valor in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)

                # Converter LOBs
                if hasattr(valor, 'read'):  # LOB object
                    try:
                        valor = valor.read()
                    except Exception as e:
                        # Só a primeira falha é logada — evita uma linha por registro
                        if not erros_lob:
                            logger.warning(f"Erro ao ler LOB: {str(e)}")
                        erros_lob += 1
                        valor = "[LOB não lido]"
                
                if isinstance(valor, bytes):
                    # Se for bytes, tenta decodificar
                    try:
                        valor = valor.decode('utf-8')
                    except:
                        valor = "[Dados binários]"
                
                if isinstance(valor, float):
                    # Formatar número com separador decimal
                    if separador == ',':
                        cell.value = str(valor).replace('.', ',')
                    else:
                        cell.value = valor
                else:
                    cell.value = valor if valor is not None else ''
        
        # Auto ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(caminho)
        if erros_lob > 1:
            logger.warning(f"{erros_lob} LOBs não puderam ser lidos em {os.path.basename(caminho)}")
    except ImportError:
        logger.error("Módulo openpyxl não instalado. Execute: pip install openpyxl")
        raise ValueError("Para exportar em XLSX, instale openpyxl: pip install openpyxl")
    except Exception as e:
        logger.error(f"Erro ao exportar para XLSX: {str(e)}")
        raise

def compactar_csvs(caminhos_arquivo, nome_zip):
    """Compacta múltiplos arquivos em um ZIP com compressão e remove os originais"""
    try:
        with zipfile.ZipFile(nome_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for arquivo in caminhos_arquivo:
                if os.path.exists(arquivo):
                    arcname = os.path.basename(arquivo)
                    zipf.write(arquivo, arcname=arcname)

        logger.info(f"ZIP criado: {nome_zip} ({len(caminhos_arquivo)} arquivo(s))")

        # Remover arquivos originais após compactação bem-sucedida
        for arquivo in caminhos_arquivo:
            try:
                if os.path.exists(arquivo):
                    os.remove(arquivo)
            except Exception as e:
                logger.warning(f"Erro ao remover arquivo {arquivo}: {str(e)}")
        
        return nome_zip
    except Exception as e:
        logger.error(f"Erro ao compactar arquivos: {str(e)}")
        raise

def salvar_apikeys_secure(arquivo, apikeys):
    """Atualiza apenas as linhas api.<nome>.apikey no secure.properties, preservando
    todo o restante do arquivo (credenciais Oracle, e-mail, MDM, etc.).

    apikeys: dict {nome_api: apikey}. APIs ausentes têm sua apikey removida do arquivo.
    Também remove chaves api.* legadas (url/header/method/params) que porventura
    tenham ficado em secure.properties — esses metadados agora vivem em config.properties.
    """
    try:
        linhas_preservadas = []
        if os.path.exists(arquivo):
            with open(arquivo, 'r', encoding='utf-8') as f:
                for line in f:
                    stripped = line.strip()
                    if stripped and not stripped.startswith('#') and '=' in stripped:
                        key = stripped.split('=', 1)[0].strip()
                        if key.startswith('api.'):
                            continue  # descarta qualquer linha de API — será reescrita
                    linhas_preservadas.append(line.rstrip('\n'))

        # Remove o cabeçalho antigo da seção de APIs, se existir, para não duplicar
        marcador = '# === APIs externas'
        linhas_preservadas = [l for l in linhas_preservadas if not l.strip().startswith(marcador)]

        # Remove linhas em branco redundantes no fim
        while linhas_preservadas and linhas_preservadas[-1].strip() == '':
            linhas_preservadas.pop()

        linhas = list(linhas_preservadas)
        apikeys_validas = {n: k for n, k in apikeys.items() if k}
        if apikeys_validas:
            linhas.append('')
            linhas.append('# === APIs externas - APIKEY (sensível) ===')
            for nome in sorted(apikeys_validas):
                linhas.append(f'api.{nome}.apikey={apikeys_validas[nome]}')

        with open(arquivo, 'w', encoding='utf-8') as f:
            f.write('\n'.join(linhas) + '\n')
        logger.info(f"APIKEYs salvas em {arquivo} ({len(apikeys_validas)} API(s))")
    except Exception as e:
        logger.error(f"Erro ao salvar APIKEYs em {arquivo}: {str(e)}")


@app.route('/')
def index():
    props = ler_properties(CONFIG_FILE)  # tabs só existem em config.properties
    tabs = {
        'solicitar_logs': props.get('tab.solicitar_logs', 'true').lower() == 'true',
        'exportar_oracle': props.get('tab.exportar_oracle', 'true').lower() == 'true',
        'requisicao_api': props.get('tab.requisicao_api', 'true').lower() == 'true',
        'configurar_pdv': props.get('tab.configurar_pdv', 'true').lower() == 'true',
        'pinpad': props.get('tab.pinpad', 'true').lower() == 'true',
        'mdm': props.get('tab.mdm', 'true').lower() == 'true',
        'configuracoes': props.get('tab.configuracoes', 'true').lower() == 'true',
    }
    return render_template('index.html', tabs=tabs)

@app.route('/solicitar-logs')
def solicitar_logs_page():
    return render_template('solicitar_logs.html', config_props=ler_config_completo())

@app.route('/exportar-oracle')
def exportar_oracle_page():
    return render_template('exportar_oracle.html', config_props=ler_config_completo())

@app.route('/configurar-pdv')
def configurar_pdv_page():
    return pagina_configurar_pdv(app, ler_config_completo, CONFIG_FILE)

@app.route('/enviar-config-pdv', methods=['POST'])
def enviar_config_pdv_route():
    return enviar_configuracao_pdv(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_ao_agente)

@app.route('/verificar-config-pdv', methods=['POST'])
def verificar_config_pdv_route():
    return verificar_configuracao_pdv(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_ao_agente)

@app.route('/relatorio-parametrizacao', methods=['POST'])
def relatorio_parametrizacao_route():
    return relatorio_parametrizacao(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_ao_agente)

@app.route('/versao-pdv', methods=['POST'])
def versao_pdv_route():
    return versao_pdv(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_ao_agente)

@app.route('/reiniciar-pdv', methods=['POST'])
def reiniciar_pdv_route():
    return reiniciar_pdv(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_ao_agente)

@app.route('/fechar-pdv', methods=['POST'])
def fechar_pdv_route():
    return fechar_pdv(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_ao_agente)

@app.route('/requisicao-api')
def requisicao_api_page():
    return pagina_requisicao_api(app, ler_config_completo, CONFIG_FILE)

@app.route('/fazer-requisicao-api', methods=['POST'])
def fazer_requisicao_api_route():
    return fazer_requisicao_api(app, ler_config_completo, CONFIG_FILE)

@app.route('/requisicao-api/gerar', methods=['POST'])
def requisicao_api_gerar_route():
    return gerar_retorno_download(app, ler_config_completo, CONFIG_FILE)

@app.route('/requisicao-api/baixar/<nome_arquivo>')
def requisicao_api_baixar_route(nome_arquivo):
    return baixar_retorno(app, nome_arquivo)

@app.route('/requisicao-api/email', methods=['POST'])
def requisicao_api_email_route():
    return enviar_retorno_email(app, ler_config_completo, CONFIG_FILE, gerar_pid, enviar_email_com_anexos)

@app.route('/mdm')
def mdm_page():
    return pagina_mdm(app, ler_config_completo, CONFIG_FILE)

@app.route('/mdm-consultar', methods=['POST'])
def mdm_consultar_route():
    return consultar_cliente_mdm(app, ler_config_completo, CONFIG_FILE)

@app.route('/mdm-cadastrar', methods=['POST'])
def mdm_cadastrar_route():
    return cadastrar_cliente_mdm(app, ler_config_completo, CONFIG_FILE)

@app.route('/mdm-atualizar', methods=['POST'])
def mdm_atualizar_route():
    return atualizar_cliente_mdm(app, ler_config_completo, CONFIG_FILE)


def converterParaArray(valor):
    if isinstance(valor, str):
        return [v.strip() for v in valor.split(',') if v.strip()]
    return valor if isinstance(valor, list) else []


# Campos do secure.properties editáveis na aba Administrador.
# sensivel=True: a tela nunca exibe o valor gravado; em branco = manter o atual.
CAMPOS_SECURE = [
    ('email_envio',             False),
    ('senha_envio',             True),
    ('oracle_host',             False),
    ('oracle_port',             False),
    ('oracle_service',          False),
    ('oracle_user',             False),
    ('oracle_password',         True),
    ('mdm_api_url',             False),
    ('mdm_api_schema',          False),
    ('mdm_api_apikey',          True),
    ('cloudflared_credentials', False),
    ('cloudflared_tunnel_name', False),
]


def _campos_secure_do_form(form):
    """Monta o dict a gravar no secure.properties a partir do formulário.

    Campos sensíveis deixados em branco viram None (mantém o valor atual), pois a
    tela os exibe vazios — salvar a string vazia apagaria a credencial.
    """
    valores = {}
    for chave, sensivel in CAMPOS_SECURE:
        campo = f'secure_{chave}'
        if campo not in form:
            continue
        valor = form.get(campo, '').strip()
        valores[chave] = None if (sensivel and not valor) else valor
    return valores

@app.route('/config', methods=['GET', 'POST'])
def config():
    props = ler_properties(CONFIG_FILE)  # apenas props editáveis

    if request.method == 'POST':
        # Abas
        for tab in ['solicitar_logs', 'exportar_oracle', 'requisicao_api', 'configurar_pdv', 'pinpad', 'mdm', 'configuracoes']:
            props[f'tab.{tab}'] = 'true' if request.form.get(f'tab_{tab}') else 'false'

        # E-mails destino (lista de destinatários — não sensível)
        emails_raw = request.form.get('emails_destino', '').replace('\n', ',')
        props['emails_destino'] = ','.join(e.strip() for e in emails_raw.split(',') if e.strip())

        # Modo de instalação
        props['modo_instalacao'] = request.form.get('modo_instalacao', 'pc')
        props['bec_loja'] = request.form.get('bec_loja', '').strip()
        props['bec_pdv']  = request.form.get('bec_pdv', '').strip()

        # PinPad porta e modo de comunicação
        pinpad_porta = request.form.get('pinpad_porta', '').strip()
        if pinpad_porta:
            props['pinpad_porta'] = pinpad_porta
        props['pinpad_modo_comunicacao'] = request.form.get('pinpad_modo_comunicacao', 'email')
        props['logs_modo_comunicacao']   = request.form.get('logs_modo_comunicacao', 'email')
        props['pdv_modo_comunicacao']    = request.form.get('pdv_modo_comunicacao', 'email')
        props['registro_modo_comunicacao'] = request.form.get('registro_modo_comunicacao', 'email')
        props['atualizacao_modo_comunicacao'] = request.form.get('atualizacao_modo_comunicacao', 'email')
        props['bec_tunnel_url']          = request.form.get('bec_tunnel_url', '').strip()
        props['pinpad_tunnel_token']     = request.form.get('pinpad_tunnel_token', '').strip()

        # Lojas e PDVs
        stores_str = request.form.get('stores', props.get('stores', ''))
        props['stores'] = stores_str
        for loja in [l.strip() for l in stores_str.split(',') if l.strip()]:
            props[f'{loja}_pdvs'] = request.form.get(f'pdvs_{loja}', props.get(f'{loja}_pdvs', ''))

        # Logs
        props['logs'] = request.form.get('logs', props.get('logs', ''))

        # Logs SP (SERVERS_EP_SP) — um por linha "nome;ip" no formulário
        logs_sp_raw = request.form.get('logs_sp', '').replace('\n', ',')
        props['logs_sp'] = ','.join(l.strip() for l in logs_sp_raw.split(',') if l.strip())

        # Parâmetros PDV
        parametros_pdv_str = request.form.get('parametros_pdv', '').strip()
        if parametros_pdv_str:
            props['PARAMETROS_PDV'] = parametros_pdv_str

        # Lojas ignoradas (Relatório e Versão)
        props['ignorar_lojas'] = ','.join(
            l.strip() for l in request.form.get('ignorar_lojas', '').split(',') if l.strip()
        )

        # Oracle – adicionar nova consulta
        nome_nova_consulta = request.form.get('oracle_query_name', '').strip()
        sql_nova_consulta  = request.form.get('oracle_query_sql', '').strip()
        if nome_nova_consulta and sql_nova_consulta:
            lista = [n.strip() for n in props.get('oracle_query_names', '').split(',') if n.strip()]
            if nome_nova_consulta not in lista:
                lista.append(nome_nova_consulta)
            props['oracle_query_names'] = ','.join(lista)
            props[f'oracle_query.{nome_nova_consulta}'] = sql_nova_consulta

        # Oracle – editar consultas existentes
        for nome in [n.strip() for n in props.get('oracle_query_names', '').split(',') if n.strip()]:
            key = f'oracle_query_sql_{nome}'
            if key in request.form:
                props[f'oracle_query.{nome}'] = request.form.get(key, '')

        # Oracle – remover consulta
        remover = request.form.get('remover_consulta', '')
        if remover:
            nomes = [n.strip() for n in props.get('oracle_query_names', '').split(',') if n.strip()]
            props['oracle_query_names'] = ','.join(n for n in nomes if n != remover)
            props.pop(f'oracle_query.{remover}', None)

        # APIs – reconstrói todas a partir do formulário (metadados no config, apikey no secure).
        # A ordem dos blocos no formulário (índices 0,1,2...) define a ordem do combobox.
        for k in [k for k in props if k.startswith('api.')]:
            props.pop(k, None)
        apikeys = {}
        indices = sorted(
            {int(c.rsplit('_', 1)[-1]) for c in request.form if c.startswith('api_nome_')}
        )
        ordem = []
        for idx in indices:
            nome = request.form.get(f'api_nome_{idx}', '').strip()
            if not nome:
                continue
            props[f'api.{nome}.url']        = request.form.get(f'api_url_{idx}', '').strip()
            props[f'api.{nome}.method']     = request.form.get(f'api_method_{idx}', 'GET').strip().upper()
            props[f'api.{nome}.params']     = request.form.get(f'api_params_{idx}', '').strip()
            props[f'api.{nome}.param_hint'] = request.form.get(f'api_param_hint_{idx}', '').strip()
            props[f'api.{nome}.body']       = request.form.get(f'api_body_{idx}', '').strip()
            # Headers é multilinha; codifica quebras como "\n" literal p/ caber em 1 linha do .properties
            headers_raw = request.form.get(f'api_headers_{idx}', '').replace('\r\n', '\n').replace('\r', '\n').strip()
            props[f'api.{nome}.headers']    = headers_raw.replace('\n', '\\n')
            props[f'api.{nome}.token_provider'] = request.form.get(f'api_token_provider_{idx}', '').strip()
            apikeys[nome] = request.form.get(f'api_apikey_{idx}', '').strip()
            ordem.append(nome)
        props['api_order'] = ','.join(ordem)

        salvar_properties(CONFIG_FILE, props)
        salvar_apikeys_secure(SECURE_FILE, apikeys)

        # Aba Administrador — só grava quando a flag está ativa, para que um POST
        # forjado não altere credenciais em instalações sem a aba.
        if props.get('admin_habilitado', 'false').strip().lower() == 'true':
            salvar_secure_campos(SECURE_FILE, _campos_secure_do_form(request.form))

        return redirect('/config?saved=1')

    # GET – preparar dados (somente props editáveis para o formulário)
    lojas = converterParaArray(props.get('stores', ''))
    pdvs_dict = {f'{loja}_pdvs': converterParaArray(props.get(f'{loja}_pdvs', '')) for loja in lojas}
    emails_destino_lista = '\n'.join(e.strip() for e in props.get('emails_destino', '').split(',') if e.strip())

    # APIs — mescla config + secure para exibir a apikey (mascarada na tela) na manutenção
    apis = obter_apis(ler_config_completo())

    # Aba Administrador — valores do secure.properties. Os sensíveis não vão para
    # a tela: só informamos se já existe algo gravado, para orientar o preenchimento.
    admin_habilitado = props.get('admin_habilitado', 'false').strip().lower() == 'true'
    secure_props = ler_properties(SECURE_FILE) if admin_habilitado else {}
    secure_valores = {}
    for chave, sensivel in CAMPOS_SECURE:
        atual = secure_props.get(chave, '')
        secure_valores[chave] = {
            'valor': '' if sensivel else atual,
            'sensivel': sensivel,
            'preenchido': bool(atual),
        }

    return render_template(
        'config.html',
        lojas=lojas,
        stores_str=','.join(lojas),
        pdvs_dict=pdvs_dict,
        logs_str=','.join(converterParaArray(props.get('logs', ''))),
        logs_sp_lista='\n'.join(converterParaArray(props.get('logs_sp', ''))),
        emails_destino_lista=emails_destino_lista,
        apis=apis,
        config_props=props,
        admin_habilitado=admin_habilitado,
        secure_valores=secure_valores,
        agentes=obter_agentes(props) if admin_habilitado else [],
        tamanho_maximo_mb=TAMANHO_MAXIMO_MB,
        emails_destino=converterParaArray(props.get('emails_destino', '')),
    )

       

@app.route('/admin/atualizar-agente', methods=['POST'])
def admin_atualizar_agente():
    """Envia o pacote de atualização ao agente selecionado (aba Administrador)."""
    props = ler_config_completo()
    if props.get('admin_habilitado', 'false').strip().lower() != 'true':
        logger.warning("Tentativa de atualizar agente com a aba Administrador desativada")
        return jsonify({'sucesso': False, 'mensagem': 'Aba Administrador não habilitada.'}), 403

    payload, status = enviar_atualizacao(props, request.form, request.files,
                                         enviar_email_com_anexos, _enfileirar_no_relay)
    return jsonify(payload), status


@app.errorhandler(413)
def _upload_muito_grande(_e):
    return jsonify({
        'sucesso': False,
        'mensagem': f'Arquivo maior que o limite de {TAMANHO_MAXIMO_MB} MB por e-mail.'
    }), 413


def _parametros_oracle_export(form):
    """Lê e valida os parâmetros comuns às ações de exportação Oracle (gerar/baixar/e-mail)."""
    consultas_str = form.get('consultas', '')
    consultas = [c.strip() for c in consultas_str.split(',') if c.strip()]
    if not consultas:
        raise ValueError("Nenhuma consulta foi selecionada")
    return {
        'consultas': consultas,
        'loja': form.get('loja', ''),
        'pdv': form.get('pdv', ''),
        'nsu': form.get('nsu', ''),
        'data': form.get('data', ''),
        'formato': form.get('formato', 'csv'),
        'separador': form.get('separador', '.'),
        'compactar': form.get('compactar', 'sim').lower() == 'sim',
    }


def gerar_arquivos_oracle(consultas, loja, pdv, nsu, data, formato, separador, compactar):
    """Gera os arquivos exportados do Oracle.

    Consultas que não retornam linhas não geram arquivo para envio/download (o arquivo
    vazio, só com cabeçalho, é descartado) — apenas aparecem no status com 0 linhas.

    Retorna (caminhos_arquivos, status_consultas):
      - caminhos_arquivos: lista de arquivos com dados prontos para anexar/baixar
        (um único ZIP se compactar=True, ou os arquivos individuais caso contrário).
        Vazia se nenhuma consulta retornou linhas.
      - status_consultas: lista de dicts {'nome', 'linhas', 'status': 'ok'|'vazio'}
        para todas as consultas selecionadas, usada para montar o relatório do e-mail.
    """
    status_consultas = []
    caminhos_com_dados = []

    for nome_consulta in consultas:
        caminho, linhas = exportar_consulta_para_csv(nome_consulta, loja, pdv, nsu, data, formato, separador)
        if linhas > 0:
            status_consultas.append({'nome': nome_consulta, 'linhas': linhas, 'status': 'ok'})
            caminhos_com_dados.append(caminho)
        else:
            status_consultas.append({'nome': nome_consulta, 'linhas': 0, 'status': 'vazio'})
            logger.info(f"Consulta {nome_consulta} não retornou linhas; arquivo descartado")
            try:
                os.remove(caminho)
            except OSError:
                pass

    if not caminhos_com_dados:
        return [], status_consultas

    if compactar:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        nome_zip = os.path.join(OUTPUT_DIR, f'OracleDB-{loja}-{pdv}-{ts}.zip')
        compactar_csvs(caminhos_com_dados, nome_zip)
        logger.info(f"Exportação compactada: {nome_zip} ({len(caminhos_com_dados)} arquivo(s))")
        return [nome_zip], status_consultas

    logger.info(f"Exportação concluída sem compactação ({len(caminhos_com_dados)} arquivo(s))")
    return caminhos_com_dados, status_consultas


@app.route('/oracle_export/gerar', methods=['POST'])
def oracle_export_gerar():
    """Gera os arquivos no servidor e retorna os nomes para download pelo navegador/webview."""
    try:
        params = _parametros_oracle_export(request.form)
    except ValueError as e:
        logger.error(str(e))
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 400

    try:
        caminhos, status_consultas = gerar_arquivos_oracle(**params)
        if not caminhos:
            logger.info("Nenhuma consulta retornou dados; nenhum arquivo gerado para download")
            return jsonify({
                'sucesso': True,
                'arquivos': [],
                'mensagem': 'Nenhuma consulta selecionada retornou dados. Nenhum arquivo foi gerado.'
            })

        registrar_execucao(ler_config_completo(), 'Exportar Oracle - Download', detalhes={
            'Consultas': ', '.join(params['consultas']),
            'Loja': params['loja'], 'PDV': params['pdv'],
            'NSU': params['nsu'], 'Data': params['data'],
            'Formato': params['formato'],
            'Arquivos': ', '.join(os.path.basename(c) for c in caminhos),
        })
        return jsonify({'sucesso': True, 'arquivos': [os.path.basename(c) for c in caminhos]})
    except Exception as e:
        logger.error(f"Erro na exportação Oracle: {str(e)}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@app.route('/oracle_export/baixar/<nome_arquivo>')
def oracle_export_baixar(nome_arquivo):
    """Serve um arquivo já gerado em OUTPUT_DIR para download."""
    nome_seguro = os.path.basename(nome_arquivo)
    caminho = os.path.realpath(os.path.join(OUTPUT_DIR, nome_seguro))
    if not caminho.startswith(os.path.realpath(OUTPUT_DIR) + os.sep) or not os.path.isfile(caminho):
        logger.error(f"Tentativa de download de arquivo inválido: {nome_arquivo}")
        return jsonify({'sucesso': False, 'mensagem': 'Arquivo não encontrado'}), 404

    logger.info(f"Download do arquivo de exportação Oracle: {nome_seguro}")
    return send_file(caminho, as_attachment=True, download_name=nome_seguro)


@app.route('/oracle_export/email', methods=['POST'])
def oracle_export_email():
    email_destino = request.form.get('email_destino', '').strip()
    if not email_destino:
        return jsonify({'sucesso': False, 'mensagem': 'Selecione um e-mail de destino'}), 400

    try:
        params = _parametros_oracle_export(request.form)
    except ValueError as e:
        logger.error(str(e))
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 400

    try:
        caminhos, status_consultas = gerar_arquivos_oracle(**params)

        props = ler_config_completo()
        pid = gerar_pid()
        corpo_txt, corpo_html = montar_email_exportacao_oracle(params, status_consultas, caminhos, pid)
        assunto = f"[Exportação Oracle][{params['loja']}][{params['pdv']}][{pid}]"

        enviar_email_com_anexos(
            props.get('email_envio', ''), props.get('senha_envio', ''),
            email_destino, assunto, corpo_txt, corpo_html, caminhos
        )
        logger.info(f"E-mail de exportação Oracle enviado para {email_destino} ({len(caminhos)} arquivo(s) anexado(s))")
        registrar_execucao(props, 'Exportar Oracle - E-mail', pid, {
            'Consultas': ', '.join(params['consultas']),
            'Loja': params['loja'], 'PDV': params['pdv'],
            'NSU': params['nsu'], 'Data': params['data'],
            'Formato': params['formato'], 'Destino': email_destino,
        })
        return jsonify({'sucesso': True, 'mensagem': f'E-mail enviado com sucesso para {email_destino}!'})
    except Exception as e:
        logger.error(f"Erro ao enviar exportação Oracle por e-mail: {str(e)}")
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


def enviar_email_com_anexos(remetente, senha, destino, assunto, corpo_texto, corpo_html, arquivos):
    """Envia um e-mail com texto/HTML alternativos e um ou mais arquivos anexados."""
    msg = MIMEMultipart('mixed')
    msg['From'] = remetente
    msg['To'] = destino
    msg['Subject'] = assunto

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(corpo_texto, 'plain', 'utf-8'))
    alt.attach(MIMEText(corpo_html, 'html', 'utf-8'))
    msg.attach(alt)

    for caminho in arquivos:
        tipo, _ = mimetypes.guess_type(caminho)
        maintype, subtype = tipo.split('/', 1) if tipo else ('application', 'octet-stream')
        with open(caminho, 'rb') as f:
            part = MIMEBase(maintype, subtype)
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(caminho)}"')
        msg.attach(part)

    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(remetente, senha)
        smtp.sendmail(remetente, destino, msg.as_string())


def montar_email_exportacao_oracle(params, status_consultas, caminhos_arquivos, pid):
    """Monta o corpo texto/HTML do e-mail de exportação Oracle, no mesmo layout visual
    usado pelo agente ao enviar arquivos de log (server_agent/agent_extrator_log.py):
    cabeçalho azul, caixas de resumo Loja/PDV/Data/PID, caixas verde/vermelha com a
    contagem de consultas com dados x sem retorno, e tabela com o status de cada consulta.
    Consultas sem linhas retornadas aparecem marcadas e não têm arquivo anexado."""
    agora_fmt = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    n_ok = sum(1 for s in status_consultas if s['status'] == 'ok')
    n_vazio = len(status_consultas) - n_ok

    linhas_html = ''
    linhas_txt = ''
    for s in status_consultas:
        if s['status'] == 'ok':
            bg_row = ''
            badge = "<span style='background:#dcfce7;color:#1a7f4b;font-size:11px;font-weight:bold;padding:3px 10px;border-radius:12px;display:inline-block'>&#10004; Incluído</span>"
        else:
            bg_row = "background:#fff5f5;"
            badge = "<span style='background:#fee2e2;color:#b91c1c;font-size:11px;font-weight:bold;padding:3px 10px;border-radius:12px;display:inline-block'>&#10006; Sem retorno</span>"

        linhas_html += f"""
        <tr style='{bg_row}'>
          <td style='padding:10px 14px;border-bottom:1px solid #e5e7eb;font-family:monospace;font-size:12px;color:#1e293b'>{s['nome']}</td>
          <td style='padding:10px 14px;border-bottom:1px solid #e5e7eb;font-size:12px;color:#6b7280;text-align:center'>{s['linhas']}</td>
          <td style='padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:center;white-space:nowrap'>{badge}</td>
        </tr>"""
        st_txt = 'OK' if s['status'] == 'ok' else 'SEM RETORNO'
        linhas_txt += f"\n  [{st_txt}] {s['nome']}  ({s['linhas']} linha(s))"

    if caminhos_arquivos:
        nomes_anexos = ', '.join(os.path.basename(c) for c in caminhos_arquivos)
        anexo_html = f"""<div style='background:#f0f9ff;border:1px solid #bae6fd;border-radius:6px;padding:12px 16px'>
      <p style='margin:0;font-size:12px;color:#0369a1'><strong>Anexo:</strong> {nomes_anexos}</p>
    </div>"""
        anexo_txt = f"Anexo: {nomes_anexos}"
    else:
        anexo_html = """<div style='background:#fff5f5;border:1px solid #fecaca;border-radius:6px;padding:12px 16px'>
      <p style='margin:0;font-size:12px;color:#b91c1c'><strong>Nenhum arquivo anexado</strong> — nenhuma consulta retornou dados.</p>
    </div>"""
        anexo_txt = "Nenhum arquivo anexado — nenhuma consulta retornou dados."

    corpo_html = f"""<!DOCTYPE html>
<html><head><meta charset='utf-8'></head>
<body style='margin:0;padding:0;background:#f3f4f6;font-family:Arial,sans-serif'>
<table width='100%' cellpadding='0' cellspacing='0' style='background:#f3f4f6;padding:24px 0'>
<tr><td align='center'>
<table width='640' cellpadding='0' cellspacing='0' style='background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)'>

  <tr><td style='background:#1e3a5f;padding:24px 28px'>
    <p style='margin:0;color:#93c5fd;font-size:12px;text-transform:uppercase;letter-spacing:1px'>Backoffice Equipe QA</p>
    <h1 style='margin:6px 0 0;color:#ffffff;font-size:20px'>Exportação de Dados Oracle</h1>
  </td></tr>

  <tr><td style='padding:20px 28px 0'>
    <table cellpadding='0' cellspacing='0' width='100%'>
      <tr>
        <td style='padding:8px 12px;background:#f8fafc;border-radius:6px;text-align:center;width:22%'>
          <p style='margin:0;font-size:11px;color:#6b7280;text-transform:uppercase'>Loja</p>
          <p style='margin:4px 0 0;font-size:18px;font-weight:bold;color:#1e3a5f'>{params['loja'] or '-'}</p>
        </td>
        <td width='8'></td>
        <td style='padding:8px 12px;background:#f8fafc;border-radius:6px;text-align:center;width:18%'>
          <p style='margin:0;font-size:11px;color:#6b7280;text-transform:uppercase'>PDV</p>
          <p style='margin:4px 0 0;font-size:18px;font-weight:bold;color:#1e3a5f'>{params['pdv'] or '-'}</p>
        </td>
        <td width='8'></td>
        <td style='padding:8px 12px;background:#f8fafc;border-radius:6px;text-align:center;width:22%'>
          <p style='margin:0;font-size:11px;color:#6b7280;text-transform:uppercase'>Data</p>
          <p style='margin:4px 0 0;font-size:13px;font-weight:bold;color:#1e3a5f;font-family:monospace'>{params['data'] or '-'}</p>
        </td>
        <td width='8'></td>
        <td style='padding:8px 12px;background:#f8fafc;border-radius:6px;text-align:center;width:26%'>
          <p style='margin:0;font-size:11px;color:#6b7280;text-transform:uppercase'>PID</p>
          <p style='margin:4px 0 0;font-size:14px;font-weight:bold;color:#1e3a5f;font-family:monospace'>{pid}</p>
        </td>
      </tr>
    </table>
  </td></tr>

  <tr><td style='padding:16px 28px'>
    <table cellpadding='0' cellspacing='0' width='100%'>
      <tr>
        <td style='background:#dcfce7;border-radius:6px;padding:10px;text-align:center;width:48%'>
          <p style='margin:0;font-size:22px;font-weight:bold;color:#1a7f4b'>{n_ok}</p>
          <p style='margin:2px 0 0;font-size:11px;color:#1a7f4b;font-weight:bold'>COM DADOS</p>
        </td>
        <td width='12'></td>
        <td style='background:#fee2e2;border-radius:6px;padding:10px;text-align:center;width:48%'>
          <p style='margin:0;font-size:22px;font-weight:bold;color:#b91c1c'>{n_vazio}</p>
          <p style='margin:2px 0 0;font-size:11px;color:#b91c1c;font-weight:bold'>SEM RETORNO</p>
        </td>
      </tr>
    </table>
  </td></tr>

  <tr><td style='padding:0 28px 8px'>
    <table width='100%' cellpadding='0' cellspacing='0' style='border:1px solid #e5e7eb;border-radius:6px;overflow:hidden;font-size:13px'>
      <thead>
        <tr style='background:#f8fafc'>
          <th style='padding:10px 14px;text-align:left;color:#6b7280;font-size:11px;text-transform:uppercase;border-bottom:1px solid #e5e7eb'>Consulta</th>
          <th style='padding:10px 14px;text-align:center;color:#6b7280;font-size:11px;text-transform:uppercase;border-bottom:1px solid #e5e7eb'>Linhas</th>
          <th style='padding:10px 14px;text-align:center;color:#6b7280;font-size:11px;text-transform:uppercase;border-bottom:1px solid #e5e7eb'>Status</th>
        </tr>
      </thead>
      <tbody>{linhas_html}
      </tbody>
    </table>
  </td></tr>

  <tr><td style='padding:8px 28px 24px'>
    {anexo_html}
  </td></tr>

  <tr><td style='padding:14px 28px;background:#f8fafc;border-top:1px solid #e5e7eb'>
    <p style='margin:0;font-size:11px;color:#9ca3af'>Gerado em {agora_fmt} &nbsp;|&nbsp; Backoffice Equipe QA</p>
  </td></tr>

</table>
</td></tr>
</table>
</body></html>"""

    corpo_txt = (
        f"Exportação de Dados Oracle\n"
        f"PID: {pid} | Loja: {params['loja']} | PDV: {params['pdv']} | Data: {params['data']}\n"
        f"Resumo: {n_ok} com dados | {n_vazio} sem retorno\n"
        f"{anexo_txt}\n"
        f"{'=' * 60}"
        + linhas_txt
    )

    return corpo_txt, corpo_html


# Assunto do e-mail -> tipo do item na fila do relay. É o que permite rotear as
# funcionalidades de Manutenção PDV sem alterar o corpo que elas já montam: o
# corpo segue igual nos dois canais, e o agente usa o mesmo handler.
_TIPOS_RELAY_POR_ASSUNTO = (
    ('[Parametrização PDV]',      'parametrizacao_pdv'),
    ('[Verificar Parametrização]', 'verificar_parametrizacao'),
    ('[Relatório Parametrização]', 'relatorio_parametrizacao'),
    ('[Status PDV]',              'status_pdv'),
    ('[Fechar PDV]',              'fechar_pdv'),
    ('[Reiniciar PDV]',           'reiniciar_pdv'),
)


def _tipo_relay_do_assunto(assunto):
    for marcador, tipo in _TIPOS_RELAY_POR_ASSUNTO:
        if marcador in assunto:
            return tipo
    return None


def _enfileirar_no_relay(props, tipo, pid, corpo, extras=None):
    """Publica um item na fila do agente no relay.

    Levanta RuntimeError em qualquer falha, para que as funcionalidades já
    existentes tratem o erro pelo mesmo `except` que usavam com o e-mail.
    """
    worker_url = props.get('bec_tunnel_url', '').rstrip('/')
    token      = props.get('pinpad_tunnel_token', '')
    bec_loja   = props.get('bec_loja', '')
    bec_pdv    = props.get('bec_pdv', '')

    if not worker_url:
        raise RuntimeError('URL do relay não configurada (Configurações → Modo de comunicação).')
    if not bec_loja or not bec_pdv:
        raise RuntimeError('Loja e PDV do BEC não configurados — é o par que identifica a fila do agente no relay.')

    # A fila é endereçada pelo par loja/pdv do AGENTE (bec_loja/bec_pdv). O alvo
    # de cada operação vai dentro do corpo, como no e-mail.
    payload = {
        'tipo':    tipo,
        'pid':     pid,
        'usuario': USUARIO_WINDOWS,
        'corpo':   corpo,
    }
    if extras:
        payload.update(extras)

    logger.info(f"[Relay] Enfileirando {tipo} PID={pid} na fila {bec_loja}/{bec_pdv}")
    status, _ = _worker_call('POST', f'{worker_url}/comando/{bec_loja}/{bec_pdv}', payload, token)
    if status not in (200, 201):
        raise RuntimeError(f'o relay recusou a solicitação (HTTP {status}). Verifique a URL e o token.')

    logger.info(f"[Relay] {tipo} enfileirada com sucesso. PID: {pid}")


# Deixa o registro de execução alcançar a fila do relay sem que o execucao.py
# precise importar este módulo de volta.
execucao.definir_canal_relay(_enfileirar_no_relay)


def enviar_ao_agente(remetente, senha, destinatario, assunto, corpo):
    """Entrega uma solicitação ao agente pelo canal configurado.

    Tem de propósito a mesma assinatura de `enviar_email_gmail`: as
    funcionalidades de Manutenção PDV recebem esta função no lugar dela e não
    precisam saber qual canal está em uso. Cai no e-mail quando a flag da
    funcionalidade está em `email` ou quando o assunto não é roteável.
    """
    tipo = _tipo_relay_do_assunto(assunto)
    if not tipo:
        return enviar_email_gmail(remetente, senha, destinatario, assunto, corpo)

    props = ler_config_completo()
    if props.get('pdv_modo_comunicacao', 'email').strip().lower() != 'tunnel':
        return enviar_email_gmail(remetente, senha, destinatario, assunto, corpo)

    # O PID já está no corpo montado pela funcionalidade — reaproveita em vez de
    # gerar outro, para a trilha de ações casar com o que a tela mostrou.
    m = re.search(r'PID:\s*(\S+)', corpo)
    pid = m.group(1) if m else gerar_pid()
    _enfileirar_no_relay(props, tipo, pid, corpo)


# Resultado das solicitacoes de log em modo download, por PID. A leitura do
# resultado no relay e destrutiva (consome), entao guardar aqui permite a tela
# consultar o status varias vezes sem perder a informacao. Some quando o arquivo
# e baixado; o processo do BEC e de vida curta, nao precisa de expiracao.
_RESULTADOS_LOG = {}


def _solicitar_logs_via_relay(props, pid, loja, pdv, logs, email_destino, data_logs, corpo, entrega='email'):
    """Enfileira a solicitação de logs no relay, em vez de enviar e-mail.

    Só o pedido deixa de usar e-mail: o agente responde com o zip dos logs por
    e-mail, como sempre. O registro da ação na trilha de usuários é feito pelo
    próprio agente ao consumir a fila — mandar aqui um [Registro Execucao] por
    e-mail anularia justamente o que este modo evita.
    """
    # Campos soltos além do `corpo`: um agente ainda na v1.2.0 os espera, pois
    # naquela versão a solicitação de log era transportada campo a campo.
    extras = {
        'destino': email_destino,
        'loja':    loja,
        'pdv':     pdv,
        'logs':    ', '.join(logs),
        'data':    data_logs,
        'entrega': entrega,
    }

    try:
        _enfileirar_no_relay(props, 'solicitacao_log', pid, corpo, extras)
    except Exception as e:
        logger.error(f"[Relay] Falha ao enfileirar solicitação de logs: {e}")
        return jsonify({'sucesso': False, 'mensagem': f'Erro ao enviar a solicitação: {e}'}), 500

    if entrega == 'download':
        mensagem = (f'Solicitação enviada! PID: {pid}. Extraindo os logs — o download '
                    f'começa automaticamente quando o agente terminar.')
    else:
        mensagem = (f'Solicitação enviada pelo relay! PID: {pid}. '
                    f'Os logs chegarão por e-mail em {email_destino}.')
    return jsonify({'sucesso': True, 'mensagem': mensagem, 'pid': pid, 'entrega': entrega})


@app.route('/solicitar/status/<pid>')
def solicitar_status(pid):
    """Diz se o agente já terminou a extração deste PID.

    A tela chama isso em intervalos até o arquivo ficar pronto. A leitura do
    resultado no relay é destrutiva, então o resultado é guardado aqui em memória
    — assim o polling pode consultar quantas vezes precisar sem perdê-lo.
    """
    pid = os.path.basename(pid)
    if pid in _RESULTADOS_LOG:
        return jsonify(_RESULTADOS_LOG[pid])

    props      = ler_config_completo()
    worker_url = props.get('bec_tunnel_url', '').rstrip('/')
    token      = props.get('pinpad_tunnel_token', '')
    if not worker_url:
        return jsonify({'pronto': False, 'erro': 'URL do relay não configurada'}), 400

    status, corpo = _worker_call('GET', f'{worker_url}/resultado/{pid}', token=token)
    if status == 204:
        return jsonify({'pronto': False})       # agente ainda trabalhando
    if status != 200:
        return jsonify({'pronto': False, 'erro': f'relay respondeu HTTP {status}'})

    try:
        dados = json.loads(corpo or b'{}')
    except Exception:
        return jsonify({'pronto': False, 'erro': 'resultado ilegível no relay'})

    resultado = {
        'pronto':   True,
        'sucesso':  bool(dados.get('sucesso')),
        'mensagem': dados.get('mensagem', ''),
        'arquivo':  dados.get('arquivo', ''),
        'tamanho':  dados.get('tamanho', 0),
        'sha256':   dados.get('sha256', ''),
        'incluidos': dados.get('incluidos'),
        'faltando':  dados.get('faltando'),
    }
    _RESULTADOS_LOG[pid] = resultado
    logger.info(f"[Relay] Resultado da solicitação {pid}: sucesso={resultado['sucesso']} "
                f"arquivo={resultado['arquivo']} ({resultado['tamanho']} bytes)")
    return jsonify(resultado)


@app.route('/solicitar/baixar/<pid>')
def solicitar_baixar(pid):
    """Baixa do relay o ZIP que o agente subiu e entrega ao navegador."""
    pid = os.path.basename(pid)
    props      = ler_config_completo()
    worker_url = props.get('bec_tunnel_url', '').rstrip('/')
    token      = props.get('pinpad_tunnel_token', '')

    info = _RESULTADOS_LOG.get(pid, {})
    nome = os.path.basename(info.get('arquivo') or f'LOG-{pid}.zip')

    status, conteudo = _worker_call('GET', f'{worker_url}/arquivo/{pid}', token=token)
    if status != 200 or not conteudo:
        logger.error(f"[Relay] Download do arquivo {pid} falhou: HTTP {status}")
        return jsonify({'sucesso': False,
                        'mensagem': f'Arquivo não disponível no relay (HTTP {status}). '
                                    f'Ele expira depois de baixado ou por tempo.'}), 404

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    caminho = os.path.join(OUTPUT_DIR, nome)
    with open(caminho, 'wb') as f:
        f.write(conteudo)
    logger.info(f"[Relay] Arquivo de logs baixado: {nome} ({len(conteudo)} bytes)")

    # Libera o espaço no bucket: o arquivo já está no disco de quem pediu.
    _worker_call('DELETE', f'{worker_url}/arquivo/{pid}', token=token)
    _RESULTADOS_LOG.pop(pid, None)

    return send_file(caminho, as_attachment=True, download_name=nome)


@app.route('/solicitar', methods=['POST'])
def solicitar():
    props = ler_config_completo()

    loja = request.form['loja']
    pdv = request.form.get('pdv', '')
    logs = request.form.getlist('logs')
    email_destino = request.form['email_destino']

    # Data dos logs (campo do formulário em yyyy-mm-dd) — vai como dd/mm/yyyy no corpo
    try:
        data_obj = datetime.strptime(request.form.get('data', '').strip(), '%Y-%m-%d')
    except ValueError:
        data_obj = datetime.now()
    if data_obj.date() > datetime.now().date():
        logger.warning(f"Solicitação recusada: data futura ({data_obj.strftime('%d/%m/%Y')})")
        return jsonify({'sucesso': False, 'mensagem': 'A data dos logs não pode ser futura.'}), 400
    data_logs = data_obj.strftime('%d/%m/%Y')

    # 'email' (padrão, histórico) ou 'download' (o agente sobe o ZIP para o R2 e o
    # BEC o baixa pelo relay). O download só existe no canal relay: sem ele o
    # agente não tem para onde subir o arquivo.
    entrega = (request.form.get('entrega') or 'email').strip().lower()
    if entrega not in ('email', 'download'):
        entrega = 'email'

    logger.info(f"Solicitação de logs [loja={loja} pdv={pdv} email={email_destino} "
                f"data={data_logs} logs={','.join(logs)} entrega={entrega}]")

    pid = gerar_pid()

    if loja == 'SERVERS_EP_SP':
        # Logs de servidores extraídos pelo Server Agent SP (sem Loja/PDV no corpo)
        assunto = f"[Solicitação Log SP] - [{pid}]"
        corpo = f"PID: {pid}\nUsuario: {USUARIO_WINDOWS}\nDestino: {email_destino}\nLogs: {','.join(logs)}\nData: {data_logs}"
    elif loja == 'server_152':
        assunto = f"[Solicitação linx-webservices] - [{pid}]"
        corpo = f"PID: {pid}\nUsuario: {USUARIO_WINDOWS}\nDestino: {email_destino}\nLoja: {loja}\nPDV: {pdv}\nLogs: linx-webservices\nData: {data_logs}"
    else:
        assunto = f"[Solicitação Log] - [{pid}]"
        corpo = f"PID: {pid}\nUsuario: {USUARIO_WINDOWS}\nDestino: {email_destino}\nLoja: {loja}\nPDV: {pdv}\nLogs: {', '.join(logs)}\nData: {data_logs}"
        if entrega == 'download':
            # O agente sobe o ZIP para o R2 em vez de enviar e-mail
            corpo += "\nEntrega: download"

    # Canal de envio. O relay atende apenas o Agente Extrator (lojas de PDV): o
    # Server Agent SP não alcança o relay da rede dele, e o linx-webservices segue
    # o fluxo de e-mail de sempre. O corpo é o mesmo nos dois canais.
    atendido_pelo_extrator = loja not in ('SERVERS_EP_SP', 'server_152')
    via_relay = atendido_pelo_extrator and props.get('logs_modo_comunicacao', 'email').strip().lower() == 'tunnel'

    if entrega == 'download' and not via_relay:
        return jsonify({'sucesso': False, 'mensagem':
                        'O download exige o modo Tunnel em Configurações → Modo de comunicação, '
                        'e não está disponível para o Server Agent SP.'}), 400

    if via_relay:
        return _solicitar_logs_via_relay(props, pid, loja, pdv, logs, email_destino,
                                         data_logs, corpo, entrega)

    try:
        remetente = props.get('email_envio', '')
        enviar_email_gmail(remetente, props.get('senha_envio', ''), remetente, assunto, corpo)
        logger.info(f"Solicitação enviada com sucesso. PID: {pid}")

        # O Server Agent SP roda em outra rede e não alimenta o fluxo de registro
        # do agente principal — a execução é registrada por e-mail à parte.
        if loja == 'SERVERS_EP_SP':
            registrar_execucao(props, 'Solicitar Logs SP', pid, {
                'Logs': ','.join(logs),
                'Data': data_logs,
                'Destino': email_destino,
            })

        return jsonify({'sucesso': True, 'mensagem': f"Solicitação enviada com sucesso! PID: {pid}", 'pid': pid})
    except Exception as e:
        logger.error(f"Erro ao enviar solicitação: {str(e)}")
        return jsonify({'sucesso': False, 'mensagem': f"Erro ao enviar solicitação: {str(e)}"}), 500
    
# Função para enviar email via Gmail SMTP
def enviar_email_gmail(remetente, senha, destinatario, assunto, corpo):

    msg = MIMEText(corpo, 'plain', 'utf-8')
    msg['Subject'] = assunto
    msg['From'] = remetente
    msg['To'] = remetente

    smtp_server = "smtp.gmail.com"
    smtp_port = 587

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(remetente, senha)
        server.sendmail(remetente, destinatario, msg.as_string())




@app.route('/api/versao')
def api_versao():
    return jsonify({'versao': __version__})


IMG_DIR = os.path.abspath(os.path.join(APP_DIR, 'img'))

PINPAD_PORTA_PADRAO = 'COM10'


def _worker_call(method, url, data=None, token=''):
    """Faz uma chamada HTTP simples ao Cloudflare Worker."""
    import urllib.request as _req
    import urllib.error   as _err
    headers = {
        'X-Token':    token,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept':     'application/json',
    }
    if data is not None:
        body    = json.dumps(data).encode('utf-8')
        headers['Content-Type'] = 'application/json'
    else:
        body = None
    req = _req.Request(url, data=body, headers=headers, method=method)
    try:
        with _req.urlopen(req, timeout=10) as resp:
            return resp.status, resp.read()
    except _err.HTTPError as e:
        return e.code, b''
    except Exception as e:
        logger.warning(f'[Worker] Erro HTTP {method} {url}: {e}')
        return 0, b''


@app.route('/pinpad')
def pinpad_page():
    props = ler_properties(CONFIG_FILE)
    porta = props.get('pinpad_porta', PINPAD_PORTA_PADRAO)
    return render_template('pinpad.html', porta=porta)

def _executar_pinpad_direto(comando, porta):
    """Executa o comando serial do robô PinPad diretamente via PowerShell."""
    ps_script = (
        f"$port=New-Object System.IO.Ports.SerialPort '{porta}',115200,None,8,one;"
        f"try{{$port.Open();Start-Sleep -m 500;$port.WriteLine('{comando}');$port.Close();"
        f"echo 'Comando [{comando}] enviado para {porta}'}}catch{{echo \"ERRO: $_\"}}"
    )
    sysroot = os.environ.get('SystemRoot', r'C:\Windows')
    ps_exe = 'powershell.exe'
    for sub in (r'SysNative\WindowsPowerShell\v1.0', r'System32\WindowsPowerShell\v1.0'):
        candidato = os.path.join(sysroot, sub, 'powershell.exe')
        if os.path.exists(candidato):
            ps_exe = candidato
            break
    resultado = subprocess.run(
        [ps_exe, '-NoProfile', '-NonInteractive', '-Command', ps_script],
        capture_output=True, text=True, timeout=10
    )
    saida = (resultado.stdout or resultado.stderr or '').strip()
    sucesso = 'ERRO' not in saida.upper() and resultado.returncode == 0
    return sucesso, saida


@app.route('/pinpad/comando', methods=['POST'])
def pinpad_comando():
    props   = ler_config_completo()
    porta   = props.get('pinpad_porta', PINPAD_PORTA_PADRAO)
    data    = request.get_json()
    comando = data.get('comando', '').strip()

    comandos_validos = {'senha', 'enter', 'limpa', 'cartao'}
    if comando not in comandos_validos:
        return jsonify({'sucesso': False, 'mensagem': f'Comando inválido: {comando}'}), 400

    modo = props.get('modo_instalacao', 'pc')

    if modo == 'pdv':
        # Execução direta na máquina do PDV
        try:
            sucesso, saida = _executar_pinpad_direto(comando, porta)
            logger.info(f"PinPad direto [{comando}] -> {porta}: {saida}")
            registrar_execucao(props, 'PinPad', detalhes={
                'Comando': comando, 'Porta': porta, 'Modo': 'direto',
                'Resultado': 'sucesso' if sucesso else 'falha',
            })
            mensagem = saida or f'Comando [{comando}] enviado para {porta}'
            return jsonify({'sucesso': sucesso, 'mensagem': mensagem})
        except subprocess.TimeoutExpired:
            msg = f'Timeout ao enviar comando para {porta}'
            logger.error(f'PinPad timeout: {msg}')
            return jsonify({'sucesso': False, 'mensagem': msg}), 500
        except Exception as e:
            logger.error(f'PinPad erro direto: {str(e)}')
            return jsonify({'sucesso': False, 'mensagem': str(e)}), 500
    elif props.get('pinpad_modo_comunicacao', 'email') == 'tunnel':
        # Comunicação via Cloudflare Workers relay — independente de máquina
        loja       = props.get('bec_loja', '')
        pdv_id     = props.get('bec_pdv', '')
        worker_url = props.get('bec_tunnel_url', '').rstrip('/')
        token      = props.get('pinpad_tunnel_token', '')

        if not loja or not pdv_id:
            return jsonify({'sucesso': False, 'mensagem': 'Configure Loja e PDV do BEC nas configurações (modo de comunicação).'}), 400
        if not worker_url:
            return jsonify({'sucesso': False, 'mensagem': 'URL do Worker não configurada. Acesse Configurações → Modo de comunicação.'}), 400

        pid = gerar_pid()
        # `usuario` é obrigatório: sem ele o agente não tem como carimbar as linhas
        # de log nem registrar a ação na trilha por usuário. `tipo` explícito para
        # não depender do default de compatibilidade do agente.
        cmd_data = {
            'tipo':    'pinpad',
            'pid':     pid,
            'usuario': USUARIO_WINDOWS,
            'comando': comando,
            'porta':   porta,
        }
        logger.info(f"[Worker] Enviando PinPad [{comando}] loja={loja} pdv={pdv_id} PID={pid}")

        status, _ = _worker_call('POST', f'{worker_url}/comando/{loja}/{pdv_id}', cmd_data, token)
        if status not in (200, 201):
            logger.error(f"[Worker] Falha ao enviar comando: HTTP {status}")
            return jsonify({'sucesso': False, 'mensagem': f'Erro ao enviar comando ao relay (HTTP {status}). Verifique a URL do Worker.'}), 500

        logger.info(f"[Worker] Comando [{comando}] enviado ao relay. PID={pid}")
        # Sem registrar_execucao aqui, ao contrário do modo direto: o comando já
        # passa pelo agente, que registra a ação na trilha ao consumi-lo. Um
        # segundo POST com o mesmo PID, além de redundante, SOBRESCREVERIA o
        # comando no relay — que guarda um único item por loja/PDV.
        return jsonify({'sucesso': True, 'mensagem': f'Comando [{comando}] enviado! O dispositivo deve responder em instantes.'})
    else:
        # Envia email para o agente extrator executar remotamente
        pid     = gerar_pid()
        assunto = f"[PinPad] - [{pid}]"
        corpo   = f"PID: {pid}\nUsuario: {USUARIO_WINDOWS}\nComando: {comando}\nPorta: {porta}"
        try:
            remetente = props.get('email_envio', '')
            enviar_email_gmail(remetente, props.get('senha_envio', ''), remetente, assunto, corpo)
            logger.info(f"PinPad via email [{comando}] PID={pid}")
            return jsonify({'sucesso': True, 'mensagem': f'Comando [{comando}] enviado! PID: {pid}'})
        except Exception as e:
            logger.error(f'PinPad erro ao enviar email: {str(e)}')
            return jsonify({'sucesso': False, 'mensagem': f'Erro ao enviar solicitação: {str(e)}'}), 500


@app.route('/sobre')
def sobre():
    md_path = os.path.join(APP_DIR, 'VERSAO_ATUAL_ARQUIVOS.md')
    conteudo_md = ''
    if os.path.exists(md_path):
        with open(md_path, 'r', encoding='utf-8') as f:
            conteudo_md = f.read()
    return render_template('sobre.html', conteudo_md=conteudo_md)


@app.route('/api/tunnel/status')
def api_tunnel_status():
    """Verifica se o Worker está acessível e configurado."""
    props      = ler_config_completo()
    worker_url = props.get('bec_tunnel_url', '').rstrip('/')
    token      = props.get('pinpad_tunnel_token', '')
    if not worker_url:
        return jsonify({'ativo': False, 'url': '', 'erro': 'URL do Worker não configurada'})
    status, _ = _worker_call('GET', f'{worker_url}/status', token=token)
    if status == 200:
        return jsonify({'ativo': True, 'url': worker_url, 'erro': ''})
    return jsonify({'ativo': False, 'url': worker_url, 'erro': f'Worker inacessível (HTTP {status})'})


if __name__ == '__main__':
    import time, webview

    def _iniciar_flask():
        app.run(debug=False, port=5000, use_reloader=False, threaded=True)

    flask_thread = threading.Thread(target=_iniciar_flask, daemon=True)
    flask_thread.start()

    time.sleep(1.5)

    # Downloads são desabilitados por padrão no pywebview; habilita para permitir
    # que o botão "Download" salve arquivos na pasta Downloads do usuário.
    webview.settings['ALLOW_DOWNLOADS'] = True

    janela = webview.create_window(
        'Backoffice Equipe QA',
        'http://localhost:5000',
        width=1366,
        height=860,
        min_size=(900, 600),
    )
    webview.start()
    logger.info("Janela fechada — encerrando aplicação")
    sys.exit(0)
